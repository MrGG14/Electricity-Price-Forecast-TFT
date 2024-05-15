"""
    Utils File for PoC Bidding Energia: functions for plotting
    Last version April 2022
"""

from calendar import month, week
import os
import sys
from datetime import datetime, timedelta
from tkinter import E
from h11 import Data
import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
from sqlalchemy import column, values
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# Access files in different directories (only during execution)
basepath = os.path.dirname(__file__)  # script directory
# basepath = os.path.abspath("") for Notebooks

sys.path.insert(1, os.path.join(basepath, "../"))  # add root directory

# Other ways of importing
from config.config_file import *


#########################################################
# Plotting parameters and objects
#########################################################

# Modify plotting parameters
mpl.rcParams["font.family"] = "Trebuchet MS"
mpl.rcParams["font.size"] = 18
mpl.rcParams["figure.figsize"] = (15, 10)
mpl.rcParams["figure.dpi"] = 72
mpl.rcParams["xtick.bottom"] = True
mpl.rcParams["ytick.left"] = True
mpl.rcParams["axes.spines.top"] = False
mpl.rcParams["axes.spines.right"] = False
mpl.rcParams["axes.spines.bottom"] = True
mpl.rcParams["axes.spines.left"] = True


def plot_TS(
    data,
    months=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
    year=1900,
    days_of_week=[1, 2, 3, 4, 5, 6, 7],
    target=FASE1_DATASET["columns_input_v0"]["col_target"],
    legend=False,
    title="",
    xlabel="Hours of the day",
    ylabel=FASE1_DATASET["columns_input_v0"]["col_target"],
):
    """function to display in the same plot different days according to filters such as: month, year, day of the week...

    Parameters
    ----------
    data : dataframe
        dataset that we want to analyze. It must have the date indexed.
    months : list, optional
        months filter, by default [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]
    year : int, optional
        years filter, by default 1900
    days_of_week : list, optional
        days_of_week filter, by default [1, 2, 3, 4, 5, 6, 7]
    target : _type_, optional
        column we want to visualize, by default FASE1_DATASET["columns_input_v0"]["col_target"]
    legend : bool, optional
        recommended for plots with less than 10 dates, by default False
    title : str, optional
        Plot title, by default ""
    xlabel : str, optional
        xlabel description, by default "Hours of the day"
    ylabel : _type_, optional
        ylabel description, by default FASE1_DATASET["columns_input_v0"]["col_target"]
    """

    data["month"] = data.index.month
    data["day_of_week"] = data.index.day_of_week + 1
    data["hour_day"] = data.index.hour
    data["year"] = data.index.year
    data["day_of_year"] = data.index.day_of_year

    fig, ax = plt.subplots()

    for date in data.loc[
        (data.hour_day == 0)
        & (data.month.isin(months))
        & (data.year > year)
        & (data.day_of_week.isin(days_of_week))
    ].index:
        date_1 = str(
            datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S") + timedelta(days=1)
        )

        ax.plot(
            data.hour_day.loc[(data.index >= date) & (data.index < date_1)],
            data[target].loc[(data.index >= date) & (data.index < date_1)],
            label=str(date).replace(" 00:00:00", ""),
        )

        if legend == True:
            ax.legend()

    ax.set_title(title)
    ax.set_xlabel(xlabel, labelpad=30)
    ax.set_ylabel(ylabel, labelpad=30)


def compare_plots_TS(
    dict_predict,
    date,
    xlabel="Hours of the day",
    ylabel="Spot Price",
    color=PLOTTING["colors"],
    linestyle=PLOTTING["linestyle"],
):
    """function to compare the predictions of different models

    Parameters
    ----------
    dict_predict : dict
        dictionary with the model name as key and the prediction as value
    date : str
        date of comparison
    """

    fig, ax = plt.subplots()

    title = str(date)[:10] + " Forecast"

    for key, value in dict_predict.items():
        print(key)

        ax.plot(
            list(range(24)),
            value,
            label=key,
            color=color[key],
            linestyle=linestyle[key],
        )

        ax.legend(fontsize="x-small")

        # title = title + key + " - "

    ax.set_title(title)
    ax.set_xlabel(xlabel, labelpad=30)
    ax.set_ylabel(ylabel, labelpad=30)


def plot_prediction_real(date, suf=""):
    """function to generate a history of plots where the predictions and the actual price will be shown

    Parameters
    ----------
    date : str
        day of the forecast
    suf : str, optional
        suffix to work with different excel, by default ""
    """

    fold_path_OUT = os.path.join(PATHS["LOCAL"]["root"], PATHS["LOCAL"]["output_fase1"])

    wb = load_workbook(os.path.join(fold_path_OUT, "dataset_output" + suf + ".xlsx"))
    sheets = wb.sheetnames

    date_0 = datetime.strptime(date, "%Y-%m-%d %H:%M:%S") - timedelta(days=2)
    date_1 = datetime.strptime(date, "%Y-%m-%d %H:%M:%S") - timedelta(days=1)

    df = pd.read_excel(os.path.join(fold_path_OUT, "dataset_output" + suf + ".xlsx"))

    df_add = df[(df["DATE"] > date_0) & (df["DATE"] <= date_1)].copy()

    dict_compare_prediction = {}

    dict_compare_prediction["price_real"] = df_add["REAL"].values

    for i in sheets:
        df = pd.read_excel(
            os.path.join(fold_path_OUT, "dataset_output" + suf + ".xlsx"), sheet_name=i
        )

        df_add = df[(df["DATE"] > date_0) & (df["DATE"] <= date_1)].copy()

        dict_compare_prediction[i] = df_add["PREDICT"].values

    compare_plots_TS(dict_compare_prediction, date_1)

    name_pdf = "Forecast_" + str(date_1)[:10] + ".pdf"

    fold_path_plot = os.path.join(PATHS["root"], PATHS["plot"])

    plt.savefig(
        os.path.join(fold_path_plot, "Fase1/Plots_Hist/", name_pdf),
        format="pdf",
        bbox_inches="tight",
    )
