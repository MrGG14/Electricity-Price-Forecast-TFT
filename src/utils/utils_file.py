"""
    Utils File for PoC Bidding Energia: general functions
    Last version April 2022
"""

# Importacion de librerias
from calendar import week
import os
import sys
from datetime import datetime, timedelta
import numpy as np
import pandas as pd
from unidecode import unidecode

# from typing import Tuple, List
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Access files in different directories (only during execution)
basepath = os.path.dirname(__file__)  # script directory
# basepath = os.path.abspath("") for Notebooks


sys.path.insert(1, os.path.join(basepath, ".."))  # add root directory

# Other ways of importing
from config.metadata import FASE1_DATASET, PATHS, DIR


def get_dataset_input_v1(date: str, ponderada: int = 1) -> pd.DataFrame:
    """
    En esta funcion se procesa el dataset original y se crea el dataset que se usa en modelacion.
    Se crea el dataset final concatenando la informacion historica y las predicciones
    las variables para las horas donde no se tiene informacion historica.

    Dichas predicciones no son calculadas con modelos nuestros sino que se extraen o bien a traves de web scrapping o con una API

    Parameters
    ----------
    date : str
        day and time of the last value to be predicted  Format: 'YYYY-MM-DD HH:MM:SS'
    ponderada : int, optional
    # TODO: pending description
        __description__ , by default 0

    Returns
    -------
    pd.DataFrame
        _description_
    """

    # Se define la carpeta de trabajo

    fold_path = os.path.join(PATHS[DIR]["root"], PATHS[DIR]["src_fase1"])

    # Se importa el excel con los datos extraidos del web scrapping y de la api.
    data = pd.read_excel(
        os.path.join(fold_path, PATHS[DIR]["dataset_input_name"])
    )

    # Se hacen calculos para ponderar la demanda teniendo en cuenta factor constante= 0.1 las 24h.
    if ponderada == 1:
        data["demanda"] = data["demanda"] * 1.1
        data["demanda_prev12"] = data["demanda_prev12"] * 1.1
        data["demanda_prev36"] = data["demanda_prev36"] * 1.1

    # Se hacen calculos para ponderar la demanda calculando el factor de interconexión medio según el día de la semana y la hora
    elif ponderada == 2:
        # Calculamos los factores
        data["factor"] = 1 + (data["interconexion_global"] / data["demanda"])
        data["factor12"] = 1 + (
            data["interconexion_global"] / data["demanda_prev12"]
        )
        data["factor36"] = 1 + (
            data["interconexion_global"] / data["demanda_prev36"]
        )

        # Creamos columnas con varibles de fechas para poder hacer el calculo por dia y hora
        data["day"] = data["fechaHora"].dt.day_name()
        data["hour"] = data["fechaHora"].dt.hour
        data["day_hour"] = data["day"].astype(str) + data["hour"].astype(str)

        # Creamos un nuevo df con el periodo de las ultimas 4 semanas (28 dias) para calcular el factor
        df = data[
            (
                data["fechaHora"]
                > (
                    pd.to_datetime(date, format="%Y/%m/%d")
                    - timedelta(days=28)
                )
            )
        ]
        # Calculamos la media de los factores para las ultimas 4 semanas + 1 dia a futuro
        factor_interconexion = (
            df[["factor", "factor12", "factor36"]]
            .groupby(df["day_hour"])
            .mean()
        )
        # Calculamos agregamos la columna al df original
        data = pd.merge(data, factor_interconexion, on="day", how="left")

        # Substituimos las columnas demanda por la misma multiplicada por el factor calculado
        data["demanda"] = data["demanda"] * data["factor_y"]
        data["demanda_prev12"] = data["demanda_prev12"] * data["factor12_y"]
        data["demanda_prev36"] = data["demanda_prev36"] * data["factor36_y"]

    # FASE1_DATASET es una variable global que tiene en un diccionario con informacion relativa a la fase 1
    # si extraemos el valor seleccionando la key col_date tenemos fechaHora, que es la columna de fecha de nuestro dataset.
    # Con esta mascara booleana estamos extrayendo los registros anteriores a la fecha date

    data = data[
        (data[FASE1_DATASET["columns_input_v0"]["col_date"]] <= date)
    ].copy()

    # A continuacion se va a realizar un proceso de concatenacion en varias variables.
    # Se explica para la variable demanda.
    # En el momento de la prediccion se utilizan ciertos datos de demanda:
    # 1. la demanda real: constituida por un historico de datos de demanda, esto es, la demanda definitiva una vez ha pasado el dia
    # 2. La prevision de demanda que se da para las X horas siguientes
    # En el momento de hacer la prediccion para el dia siguiente no se contara con la demanda real sino con la estimacion.
    # De esta forma, si se quiere dar una prediccion realista, se tendra que concatenar la serie de demanda real con la prevista y utilizar las dos informaciones

    # Para hacer esta concatenacion se definen varias fechas, estas fechas mas adelante se utilizaran como limites de diferentes tramos:
    # Se define la variable end_actual_0, cuyo valor son dos dias anteriores a la fecha date
    end_actual_0 = str(
        datetime.strptime(date, "%Y-%m-%d %H:%M:%S") - timedelta(days=2)
    )

    # Se define la variable end_actual_date, cuyo valor son 1 dia anterior a la fecha date
    end_actual_date = str(
        datetime.strptime(date, "%Y-%m-%d %H:%M:%S") - timedelta(days=1)
    )

    # Se define la variable prev_12_date, cuyo valor son 13 horas antes de la fecha date
    prev_12_date = str(
        datetime.strptime(end_actual_date, "%Y-%m-%d %H:%M:%S")
        - timedelta(hours=13)
    )

    # A continuacion, se definen una serie de listas con los nombres de las columnas que se usaran para los diferentes tramos
    real_columns = [
        FASE1_DATASET["columns_input_v0"]["col_date"],
        FASE1_DATASET["columns_input_v0"]["col_target"],
        FASE1_DATASET["columns_input_v0"]["col_demand"],
        FASE1_DATASET["columns_input_v0"]["col_co2"],
        FASE1_DATASET["columns_input_v0"]["col_gas"],
        FASE1_DATASET["columns_input_v0"]["col_eolic"],
        FASE1_DATASET["columns_input_v0"]["col_solar"],
        FASE1_DATASET["columns_input_v0"]["col_demanda_residual"],
    ]

    pred_0 = [
        FASE1_DATASET["columns_input_v0"]["col_date"],
        FASE1_DATASET["columns_input_v0"]["col_target"],
        FASE1_DATASET["columns_input_v0"]["col_demand"],
        FASE1_DATASET["columns_input_v0"]["col_co2_prev24"],
        FASE1_DATASET["columns_input_v0"]["col_gas_prev24"],
        FASE1_DATASET["columns_input_v0"]["col_eolic"],
        FASE1_DATASET["columns_input_v0"]["col_solar"],
        FASE1_DATASET["columns_input_v0"]["col_demanda_residual"],
    ]
    pred_1 = [
        FASE1_DATASET["columns_input_v0"]["col_date"],
        FASE1_DATASET["columns_input_v0"]["col_target"],
        FASE1_DATASET["columns_input_v0"]["col_demand_prev12"],
        FASE1_DATASET["columns_input_v0"]["col_co2_prev24"],
        FASE1_DATASET["columns_input_v0"]["col_gas_prev24"],
        FASE1_DATASET["columns_input_v0"]["col_eolic_prev12"],
        FASE1_DATASET["columns_input_v0"]["col_solar_prev12"],
        FASE1_DATASET["columns_input_v0"]["col_demanda_residual"],
    ]
    pred_2 = [
        FASE1_DATASET["columns_input_v0"]["col_date"],
        FASE1_DATASET["columns_input_v0"]["col_target"],
        FASE1_DATASET["columns_input_v0"]["col_demand_prev36"],
        FASE1_DATASET["columns_input_v0"]["col_co2_prev48"],
        FASE1_DATASET["columns_input_v0"]["col_gas_prev48"],
        FASE1_DATASET["columns_input_v0"]["col_eolic_prev36"],
        FASE1_DATASET["columns_input_v0"]["col_solar_prev36"],
        FASE1_DATASET["columns_input_v0"]["col_demanda_residual"],
    ]

    # La logica a seguir para realizar la concatenacion es la siguiente:
    # Una vez definido date, y los tramos,  miramos que informacion vamos a incluir para cada tramo
    # Para el tramo mas lejano al momento actual se utilizara solo informacion pasada historica
    # Si por el contrario estamos en un tramo mas cercano a la fecha date se utilizaran previsiones.
    # Esto se hace para mantener la maxima realidad posible dentro de las predicciones. Al hacer simulaciones se podrian coger simplemente datos pasados
    # historicos, sin embargo, esto no representaria la realidad del dia a dia de la prediccion y podria llevar a resultados erroneos
    # A continuacion se definen los diferentes tramos y los valores que se van asignando en cada uno de ellos

    # Se define el dataset real_data_0 donde se utilizan las columnas de pred0 cuyas fechas estan entre end_actual_0 y prev_12_date,
    # es decir, estan entre dos dias antes y 12 h antes del dia que se quiere predecir
    real_data_0 = data.loc[
        (data[FASE1_DATASET["columns_input_v0"]["col_date"]] > end_actual_0)
        & (
            data[FASE1_DATASET["columns_input_v0"]["col_date"]] <= prev_12_date
        ),
        pred_0,
    ]

    # En el caso de que exista una columna entera vacia (no existe prevision en este tramo), se rellena con la informacion historica real
    for i, j in zip(pred_0[1:], real_columns[1:]):
        if real_data_0[i].sum() == 0:
            real_data_0[i] = data.loc[
                (
                    data[FASE1_DATASET["columns_input_v0"]["col_date"]]
                    > end_actual_0
                )
                & (
                    data[FASE1_DATASET["columns_input_v0"]["col_date"]]
                    <= prev_12_date
                ),
                j,
            ].values

    # Repetimos con la misma logica pero cambiando el tramo
    real_data_1 = data.loc[
        (data[FASE1_DATASET["columns_input_v0"]["col_date"]] > prev_12_date)
        & (
            data[FASE1_DATASET["columns_input_v0"]["col_date"]]
            <= end_actual_date
        ),
        pred_1,
    ]

    # Los pasos siguientes son identicos al anterior pero cambiando los tramos y la informacion que se mete en cada tramo.
    for i, j in zip(pred_1[1:], real_columns[1:]):
        if real_data_1[i].sum() == 0:
            real_data_1[i] = data.loc[
                (
                    data[FASE1_DATASET["columns_input_v0"]["col_date"]]
                    > prev_12_date
                )
                & (
                    data[FASE1_DATASET["columns_input_v0"]["col_date"]]
                    <= end_actual_date
                ),
                j,
            ].values

    real_data_2 = data.loc[
        (
            data[FASE1_DATASET["columns_input_v0"]["col_date"]]
            > end_actual_date
        ),
        pred_2,
    ]

    for i, j in zip(pred_2[1:], real_columns[1:]):
        if real_data_2[i].sum() == 0:
            real_data_2[i] = data.loc[
                (
                    data[FASE1_DATASET["columns_input_v0"]["col_date"]]
                    > end_actual_date
                ),
                j,
            ].values

    # Una vez extraidos los dataset con la informacion de los distintos tramos se renombran sus columnas unificandolas
    real_data_0.columns = real_columns
    real_data_1.columns = real_columns
    real_data_2.columns = real_columns

    # data_end contiene toda la informacion historica de las columnas anterior a la fecha en la que empezamos a definir los tramos.
    # Por eso toda la informacion anterior a la fecha end_actual_0 (fecha donde empieza el primer tramo) es informacion historica
    data_end = data.loc[
        (data[FASE1_DATASET["columns_input_v0"]["col_date"]] <= end_actual_0),
        real_columns,
    ]

    # Se introducen todos los dataframes en una lista
    dataframes = [data_end, real_data_0, real_data_1, real_data_2]

    # Se concatenan todos los dataframes en un unico df
    df = pd.concat(dataframes)

    # Se crea la variable Rampa
    df["rampa"] = (
        df[FASE1_DATASET["columns_input_v0"]["col_demanda_residual"]].shift(1)
        - df[FASE1_DATASET["columns_input_v0"]["col_demanda_residual"]]
    )
    df = df[1:]

    return df


def select_data_length(data, end_date, df_length=24):
    """select specific length of the time series

    Parameters
    ----------
    nombre : dataFrame
        dataframe to be analyzed
    end_date : str
        End date of df.  Format: 'YYYY-MM-DD HH:MM:SS'
    df_length : int
       df sizes in weeks (recommended 24 weeks)

    Returns
    -------
    dataFrame
    """

    # Se empieza definiendo la fecha de inicio de dataset, esta es la fecha final menos el numero de semanas seleccionada como parametro inicialemnte
    start_date = str(
        datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
        - timedelta(weeks=df_length)
    )

    # Se filtra el dataset que pasamos como parametro para que este contenido entre la fecha inicial y final definida
    data = data[
        (data[FASE1_DATASET["columns_input_v0"]["col_date"]] > start_date)
        & (data[FASE1_DATASET["columns_input_v0"]["col_date"]] <= end_date)
    ].copy()

    return data


def clean_TS(data, end_date, df_length=24):
    """Generates a TS avoiding possible errors in the dates entered from excel

    Parameters
    ----------
    data : dataFrame
        data frame to be cleaned
    end_date : str
        End date of df.  Format: 'YYYY-MM-DD HH:MM:SS'
    df_length : int
       df sizes in weeks (recommended 24 weeks)

    Returns
    -------
    dataFrameSS
        df indexed to TS dates
    """

    # Se define start_date como una fecha 6 meses anterior a end_date
    start_date = str(
        datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
        - timedelta(weeks=df_length)
        + timedelta(hours=1)
    )

    # Se define un rango de fechas a nivel horario desde start_date a end_date
    TS = pd.date_range(
        start=start_date, end=end_date, freq="H", tz="Africa/Ceuta"
    ).to_list()

    # Se introduce dentro de data la variable fechaTS en la columna posicion 2
    data.insert(1, "fechaTS", TS, True)

    # Se eliminan los duplicados de la columna fecha
    data.drop_duplicates(
        subset=FASE1_DATASET["columns_input_v0"]["col_date"], inplace=True
    )

    # Se elimina la columna fecha original del dataset
    data.drop(
        FASE1_DATASET["columns_input_v0"]["col_date"], axis=1, inplace=True
    )

    # Se renombra la columna TS creada anteriormente con el nombre de la anterior columna fecha
    data.rename(
        columns={"fechaTS": FASE1_DATASET["columns_input_v0"]["col_date"]},
        inplace=True,
    )

    # Se elimina un typo que se encuentra en las ultimas 6 posiciones de la columna fecha
    data[FASE1_DATASET["columns_input_v0"]["col_date"]] = data[
        FASE1_DATASET["columns_input_v0"]["col_date"]
    ].apply(lambda x: str(x)[:-6])

    # Se convierte la columna fecha a tipo datetime
    data[FASE1_DATASET["columns_input_v0"]["col_date"]] = data[
        FASE1_DATASET["columns_input_v0"]["col_date"]
    ].apply(lambda x: datetime.strptime(x, "%Y-%m-%d %H:%M:%S"))

    # Se introduce la columna fecha como indice
    data = data.set_index(FASE1_DATASET["columns_input_v0"]["col_date"])
    data = data.asfreq("H")
    data = data.sort_index()

    cont = 0

    # Finalmente tratamos los valores faltantes y les imputamos un valor.
    # Los puntos donde la demanda es un valor nan son imputados como el valor medio entre el punto anterior y siguiente
    for i in range(
        len(data.loc[data.demanda.isna()]["precio_spot"].to_list())
    ):
        # NOTE: Si hay nulos concatenados que pasa?
        i = i - cont

        h_1 = data.loc[
            data.index
            == (data.loc[data.demanda.isna()].index[i]) - timedelta(hours=1)
        ].values

        h_2 = data.loc[
            data.index
            == (data.loc[data.demanda.isna()].index[i]) + timedelta(hours=1)
        ].values

        data.loc[data.index == (data.loc[data.demanda.isna()].index[i])] = (
            np.mean(np.array([h_1, h_2]), axis=0)
        )

        cont += 1

    return data


def scale_num_cols(data):
    """Standard scaling of numerical variables

    Parameters
    ----------
    data : dataFrame
        df to be used in the model
    target : str
        target column name

    Returns
    -------
    dataFrame
        df with scaled variables
    """
    datos_train = data.drop(
        FASE1_DATASET["columns_input_v0"]["col_target"], axis=1
    )
    numerical_cols = [
        cname
        for cname in datos_train.columns
        if datos_train[cname].dtype in ["int64", "float64"]
    ]

    scaler = StandardScaler()
    model = scaler.fit_transform(datos_train[numerical_cols])

    df = pd.DataFrame(model, index=data.index, columns=[numerical_cols])
    data[numerical_cols] = df[numerical_cols]

    return data


def apply_feature_engineering(data):
    """generate new columns

    Parameters
    ----------
    data : dataFrame
        data frame to add new columns

    Returns
    -------
    _type_
        dataFrame
    """

    # Si la produccion solar es inferior a 0 se le imputa valor 1
    data["prod_solar"].loc[data["prod_solar"] <= 0] = 1

    # Se definen variables exogenas
    exo_log = ["demanda", "prod_eolica", "prod_solar"]

    # Se calcula el loaritmo de estas variables
    for i in exo_log:
        data["log_" + i] = np.log(data[i])

    data["prod_eolica_rate"] = data["prod_eolica"] / data["demanda"]
    data["prod_solar_rate"] = data["prod_solar"] / data["demanda"]
    data["prod_eolica_solar_rate"] = (
        data["prod_eolica"] + data["prod_solar"]
    ) / data["demanda"]

    # Se crea la variable bool_we, variable binaria que indica si un dia es fin de semana o no
    data["bool_we"] = data.index.weekday
    data["bool_we"] = data["bool_we"].apply(lambda x: 1 if x > 5 else 0)

    return data


def get_dataset_output(date, data, prediction_dict, name, suf=""):
    """updates the output history and obtains MAPE from the previous day's prediction

    Parameters
    ----------
    date : str
        day and time of the last value to be predicted  Format: 'YYYY-MM-DD HH:MM:SS'
    prediction_dict : dictionary
        dictionary where the keys are the names of the models and the values are the predictions
    """
    fold_path = os.path.join(
        PATHS["LOCAL"]["root"], PATHS["LOCAL"]["src_fase1"]
    )
    fold_path_OUT = os.path.join(
        PATHS["LOCAL"]["root"], PATHS["LOCAL"]["output_fase1"]
    )

    wb = load_workbook(
        os.path.join(fold_path_OUT, "dataset_output" + suf + ".xlsx")
    )
    # wb = load_workbook(os.path.join(fold_path_OUT, name))
    sheets = wb.sheetnames

    for i, s in zip(range(len(sheets)), sheets):
        ws = wb[sheets[i]]
        df = pd.DataFrame(
            ws.values,
            columns=[
                "DATE",
                "PREDICT",
                "REAL",
                "PERCENT_ERROR",
                "DIFF",
                "MAPE",
            ],
        )
        df.drop(index=df.index[0], axis=0, inplace=True)
        ws.delete_rows(
            len(df["DATE"].to_list()) - 22, len(df["DATE"].to_list()) + 1
        )
        df_add = df[
            (
                df["DATE"]
                > datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
                - timedelta(days=2)
            )
            & (
                df["DATE"]
                <= datetime.strptime(date, "%Y-%m-%d %H:%M:%S")
                - timedelta(days=1)
            )
        ].copy()

        df_ext_real = data.copy()
        yesterday = datetime.strptime(date, "%Y-%m-%d %H:%M:%S") - timedelta(
            days=2
        )
        yesterday_23 = datetime.strptime(
            date, "%Y-%m-%d %H:%M:%S"
        ) - timedelta(days=1)
        # df_ext_real.drop(index=df.index[0], axis=0, inplace=True)
        df_ext_real_price = df_ext_real.loc[
            (df_ext_real.index > yesterday)
            & (df_ext_real.index <= yesterday_23),
            [FASE1_DATASET["columns_input_v0"]["col_target"]],
        ]

        df_add["REAL"] = df_ext_real_price.values  ###.values
        df_add["PERCENT_ERROR"] = (
            abs(df_add["PREDICT"] - df_add["REAL"]) / df_add["REAL"]
        )
        df_add["DIFF"] = df_add["PREDICT"] - df_add["REAL"]
        df_add["MAPE"] = np.mean(df_add["PERCENT_ERROR"])

        print(ws)

        for r in dataframe_to_rows(df_add, index=False, header=False):
            ws.append(r)

        # df_p = prediction_dict[s][1]
        # df_p['DATE'] = df_p['DATE'].astype('str')

        # ML:
        # for r in dataframe_to_rows(prediction_dict[s][0], index=False, header=False):
        #     ws.append(r)
        # ARIMAS:
        for r in dataframe_to_rows(
            prediction_dict[s], index=False, header=False
        ):
            ws.append(r)

    wb.save(os.path.join(fold_path_OUT, "dataset_output" + suf + ".xlsx"))
    # wb.save(os.path.join(fold_path_OUT, name))


def load_xlsx(path):
    """Makes sure index date for data_set_input_v0 is correctly setting

    Parameters
    ----------
    fold_path : Path to xlsx

    Returns
    -------
    df
        dataset_input_v0
    """
    df = pd.read_excel(os.path.join(path))
    fechaHora = FASE1_DATASET["columns_input_v0"]["col_date"]
    start_daterange = df[fechaHora][0].date()
    index_date = pd.DataFrame(
        pd.date_range(
            start=str(start_daterange),
            periods=len(df),
            freq="H",
            tz="Africa/Ceuta",
        )
    )

    index_date[0] = index_date[0].apply(lambda x: str(x)[:-6])
    index_date[0] = index_date[0].apply(
        lambda x: datetime.strptime(x, "%Y-%m-%d %H:%M:%S")
    )

    indexed_old_input = df.reset_index(drop=True).copy()
    indexed_old_input[fechaHora] = index_date[0]

    return indexed_old_input


def clean_string(
    s: str, lower_case: bool = False, space_case: bool = False
) -> str:
    """Clean a given string with ASCII-code.
    Parameters
    ----------
    s : str
        String to be normalized sith ASCII characters
    lower_case : bool, optional
        Whether to lower case the string, by default False
    space_case : bool, optional
        Whether to replace spaces by underscores, by default False
    Returns
    -------
    str
        Normalized string.
    """
    if type(s) == str:
        s = s.lower() if lower_case else s
        s = s.replace("ñ", "n")
        s = s.replace("Ñ", "N")
        s = s.replace(" ", "_") if space_case else s
        s = s.replace(".", "")
        s = unidecode(string=s, errors="replace", replace_str="_")
    return s


def tope_gas_transform(data):
    """Transform gas_price based on roof gas_price (tope gas) established by spain goverment.
        Roof gas price is €40 until 31Dec2022 with a posterior increase of €5 per month up to €70
    Parameters
    ----------
    data : dataframe input with a price_gas column

    Returns
    -------
    DataFrame
        dataframe with updated gas price column
    """
    dates = [
        ["2022-06-15 00:00:00", "2022-12-31 23:00:00"],
        ["2023-01-01 00:00:00", "2023-01-31 23:00:00"],
        ["2023-02-01 00:00:00", "2023-02-28 23:00:00"],
        ["2023-03-01 00:00:00", "2023-03-31 23:00:00"],
        ["2023-04-01 00:00:00", "2023-04-30 23:00:00"],
        ["2023-05-01 00:00:00", "2023-05-31 23:00:00"],
        ["2023-06-01 00:00:00"],
    ]

    values = [40, 45, 50, 55, 60, 65, 70]
    for date, value in zip(dates, values):
        if date[0] != "2023-06-01 00:00:00":
            data[FASE1_DATASET["columns_input_v0"]["col_gas"]].loc[
                (data.index >= date[0])
                & (data.index <= date[1])
                & (data[FASE1_DATASET["columns_input_v0"]["col_gas"]] > value)
            ] = value

        else:
            data[FASE1_DATASET["columns_input_v0"]["col_gas"]].loc[
                (data.index >= date[0])
                & (data[FASE1_DATASET["columns_input_v0"]["col_gas"]] > value)
            ] = value

    return data
